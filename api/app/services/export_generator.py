"""Serialise export result sets and store them.

Takes the columns and rows a report produced (see
``app.services.export_source``) and writes CSV, XLSX, or PDF to the configured
object storage backend via the fsspec abstraction in ``app.storage``.

Fetching is deliberately not done here. It used to be: each generator ran its
own ``SELECT * FROM marts.<table>`` against a warehouse session, which meant the
format decided what the export could read. Keeping serialisation separate is
what lets one report run against a warehouse connection and another against the
operations database.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from typing import Any

import structlog

from app.storage import get_filesystem, get_storage_path

logger = structlog.get_logger(__name__)

VALID_FORMATS = frozenset({"csv", "xlsx", "pdf"})

# Excel refuses a sheet name longer than this, or one containing []:*?/\.
_MAX_SHEET_NAME = 31
_UNSAFE_FILENAME = re.compile(r"[^A-Za-z0-9._-]+")


def _slug(name: str) -> str:
    """Return a filesystem-safe stem for a report name."""
    cleaned = _UNSAFE_FILENAME.sub("_", name).strip("._-")
    return (cleaned or "export")[:80]


def _cell(value: Any) -> str:  # noqa: ANN401 — a result-set cell is whatever the driver returned
    """Render a value for a text-based format."""
    return "" if value is None else str(value)


def to_csv(columns: list[str], rows: list[list[Any]]) -> bytes:
    """Serialise a result set as UTF-8 CSV with a BOM.

    The BOM is what makes Excel open a UTF-8 CSV without mangling accented
    characters; every other consumer skips it.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow(columns)
    writer.writerows([[_cell(v) for v in row] for row in rows])
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def to_xlsx(columns: list[str], rows: list[list[Any]], sheet_name: str = "Report") -> bytes:
    """Serialise a result set as an XLSX workbook."""
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=_UNSAFE_FILENAME.sub(" ", sheet_name)[:_MAX_SHEET_NAME] or "Report")
    ws.append(columns)
    for row in rows:
        # openpyxl writes date/datetime natively but chokes on anything exotic;
        # export_source has already coerced those to strings.
        ws.append([v if isinstance(v, (int, float, str, type(None))) else str(v) for v in row])

    # Width is set from the header alone: measuring every cell means holding the
    # whole result set to compute a cosmetic value.
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(col)) + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(columns: list[str], rows: list[list[Any]], title: str) -> bytes:
    """Serialise a result set as a paginated PDF table.

    A PDF is a fixed-width medium, so this is a readable summary rather than a
    faithful dump: columns beyond what fits on a landscape page are dropped, and
    the footer says so. CSV or XLSX is the format for complete data.
    """
    from reportlab.lib import colors  # noqa: PLC0415
    from reportlab.lib.pagesizes import landscape, letter  # noqa: PLC0415
    from reportlab.lib.styles import getSampleStyleSheet  # noqa: PLC0415
    from reportlab.lib.units import inch  # noqa: PLC0415
    from reportlab.platypus import (  # noqa: PLC0415
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    page_width = landscape(letter)[0] - inch
    max_columns = max(1, int(page_width // 72))
    shown = columns[:max_columns]
    dropped = len(columns) - len(shown)

    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        title=title,
        leftMargin=inch / 2,
        rightMargin=inch / 2,
        topMargin=inch / 2,
        bottomMargin=inch / 2,
    )

    generated = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
    note = f"{len(rows):,} rows · generated {generated}"
    if dropped:
        note += f" · {dropped} further column{'s' if dropped > 1 else ''} omitted, see the CSV"

    data = [shown] + [[_cell(v)[:40] for v in row[:max_columns]] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    doc.build(
        [
            Paragraph(title, styles["Heading2"]),
            Paragraph(note, styles["Normal"]),
            Spacer(1, 10),
            table,
        ]
    )
    return buf.getvalue()


def serialise(
    columns: list[str],
    rows: list[list[Any]],
    fmt: str,
    name: str,
) -> tuple[bytes, str]:
    """Render a result set in the requested format.

    Returns:
        Tuple of (file bytes, suggested filename).

    Raises:
        ValueError: If fmt is not one of VALID_FORMATS.
    """
    if fmt not in VALID_FORMATS:
        raise ValueError(f"Unsupported export format '{fmt}'.")

    stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    filename = f"{_slug(name)}_{stamp}.{fmt}"

    if fmt == "csv":
        content = to_csv(columns, rows)
    elif fmt == "xlsx":
        content = to_xlsx(columns, rows, sheet_name=name)
    else:
        content = to_pdf(columns, rows, title=name)

    logger.info("export.serialised", format=fmt, rows=len(rows), bytes=len(content))
    return content, filename


def store_result(job_id: int, content: bytes, filename: str) -> str:
    """Write export content to object storage and return its full path."""
    fs = get_filesystem()
    full_path = get_storage_path(f"exports/job_{job_id}/{filename}")

    parent = full_path.rsplit("/", 1)[0]
    fs.makedirs(parent, exist_ok=True)
    with fs.open(full_path, "wb") as f:
        f.write(content)

    logger.info("export.stored", job_id=job_id, path=full_path, bytes=len(content))
    return full_path


def delete_result(path: str) -> None:
    """Remove a stored export, ignoring one that is already gone.

    Retention must not stall on a file someone deleted by hand, or on a storage
    backend that is briefly unreachable — the caller clears the row either way.
    """
    try:
        fs = get_filesystem()
        if fs.exists(path):
            fs.rm(path)
    except Exception as exc:  # noqa: BLE001
        logger.warning("export.delete_failed", path=path, error=str(exc))
